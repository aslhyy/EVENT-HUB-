"""
Views para eventos.
"""
import logging

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticatedOrReadOnly, IsAuthenticated
from django.utils import timezone
from django.db.models import Q, Count, Sum, F
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from core.emails import EmailService

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse

from .models import Category, Venue, Event
from .serializers import (
    CategorySerializer, CategoryDetailSerializer,
    VenueSerializer, VenueDetailSerializer,
    EventListSerializer, EventDetailSerializer,
    EventCreateUpdateSerializer, EventStatisticsSerializer
)
from .filters import EventFilter
from core.permissions import IsEventOrganizer, IsAdminOrReadOnly

logger = logging.getLogger(__name__)


class CategoryViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de categorías.
    
    list: Listar todas las categorías
    retrieve: Obtener detalle de una categoría
    create: Crear nueva categoría (solo admin)
    update: Actualizar categoría (solo admin)
    destroy: Eliminar categoría (solo admin)
    """
    queryset = Category.objects.all()
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at', 'events_count']

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return CategoryDetailSerializer
        return CategorySerializer

    @action(detail=True, methods=['get'])
    def events(self, request, pk=None):
        """
        Obtener todos los eventos de una categoría.
        """
        category = self.get_object()
        events = category.events.filter(status='published')
        serializer = EventListSerializer(events, many=True)
        return Response(serializer.data)


class VenueViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de lugares.
    
    list: Listar todos los lugares
    retrieve: Obtener detalle de un lugar
    create: Crear nuevo lugar
    update: Actualizar lugar
    destroy: Eliminar lugar
    """
    queryset = Venue.objects.all()
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [SearchFilter, OrderingFilter, DjangoFilterBackend]
    search_fields = ['name', 'city', 'address']
    ordering_fields = ['name', 'city', 'capacity', 'created_at']
    filterset_fields = ['city', 'state', 'country']

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return VenueDetailSerializer
        return VenueSerializer

    @action(detail=True, methods=['get'])
    def events(self, request, pk=None):
        """
        Obtener todos los eventos de un lugar.
        """
        venue = self.get_object()
        events = venue.events.filter(status='published')
        serializer = EventListSerializer(events, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def nearby(self, request):
        """
        Buscar lugares cercanos (requiere lat y lon en query params).
        """
        lat = request.query_params.get('lat')
        lon = request.query_params.get('lon')
        
        if not lat or not lon:
            return Response(
                {'error': 'Se requieren parámetros lat y lon'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Aquí podrías implementar búsqueda geográfica real
        # Por ahora retornamos todos los venues
        venues = self.queryset.all()[:10]
        serializer = self.get_serializer(venues, many=True)
        return Response(serializer.data)


class EventViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de eventos.
    
    list: Listar todos los eventos
    retrieve: Obtener detalle de un evento
    create: Crear nuevo evento
    update: Actualizar evento
    destroy: Eliminar evento
    publish: Publicar evento
    unpublish: Despublicar evento
    cancel: Cancelar evento
    statistics: Obtener estadísticas del evento
    """
    queryset = Event.objects.select_related(
        'category', 'venue', 'organizer'
    ).all()
    permission_classes = [IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = EventFilter
    search_fields = ['title', 'description', 'tags']
    ordering_fields = ['start_date', 'created_at', 'views_count', 'capacity']

    def get_serializer_class(self):
        if self.action == 'list':
            return EventListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return EventCreateUpdateSerializer
        elif self.action == 'statistics':
            return EventStatisticsSerializer
        return EventDetailSerializer

    def get_permissions(self):
        """Permisos según la acción."""
        if self.action in ['update', 'partial_update', 'destroy', 'publish', 'unpublish', 'cancel']:
            return [IsAuthenticated(), IsEventOrganizer()]
        return super().get_permissions()

    def retrieve(self, request, *args, **kwargs):
        """Incrementar contador de vistas al ver detalle."""
        instance = self.get_object()
        instance.views_count = F('views_count') + 1
        instance.save(update_fields=['views_count'])
        instance.refresh_from_db()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def publish(self, request, pk=None):
        """
        Publicar un evento.
        """
        event = self.get_object()
        
        if event.status == 'published':
            return Response(
                {'error': 'El evento ya está publicado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        event.status = 'published'
        event.save()
        
        serializer = self.get_serializer(event)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def unpublish(self, request, pk=None):
        """
        Despublicar un evento.
        """
        event = self.get_object()
        
        if event.status != 'published':
            return Response(
                {'error': 'El evento no está publicado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        event.status = 'draft'
        event.save()
        
        serializer = self.get_serializer(event)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """
        Cancelar un evento.
        """
        event = self.get_object()
        
        if event.status == 'cancelled':
            return Response(
                {'error': 'El evento ya está cancelado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        event.status = 'cancelled'
        event.save()

       # Notificar cancelación por email
        from core.emails import EmailService
        import logging
        logger = logging.getLogger(__name__)
        
        try:
            EmailService.send_event_cancelled(event)
            logger.info(f"Emails de cancelación enviados para {event.title}")
        except Exception as e:
            logger.error(f"Error enviando emails de cancelación: {e}")
        
        serializer = self.get_serializer(event)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """
        Obtener estadísticas del evento.
        """
        event = self.get_object()
        
        # Importar aquí para evitar circular import
        from apps.tickets.models import Ticket
        from apps.attendees.models import CheckInLog
        
        tickets = Ticket.objects.filter(
            ticket_type__event=event,
            status='active'
        )
        
        total_revenue = sum(
            ticket.ticket_type.price for ticket in tickets
        )
        
        checked_in = CheckInLog.objects.filter(
            attendee__ticket__ticket_type__event=event
        ).values('attendee').distinct().count()
        
        stats = {
            'total_capacity': event.capacity,
            'tickets_sold': event.tickets_sold,
            'tickets_available': event.tickets_available,
            'revenue': total_revenue,
            'attendees_checked_in': checked_in,
            'conversion_rate': (event.tickets_sold / event.capacity * 100) if event.capacity > 0 else 0
        }
        
        serializer = EventStatisticsSerializer(stats)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def upcoming(self, request):
        """
        Obtener eventos próximos.
        """
        events = self.queryset.filter(
            status='published',
            start_date__gte=timezone.now()
        ).order_by('start_date')[:10]
        
        serializer = EventListSerializer(events, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def featured(self, request):
        """
        Obtener eventos destacados (más vistos).
        """
        events = self.queryset.filter(
            status='published',
            start_date__gte=timezone.now()
        ).order_by('-views_count')[:10]
        
        serializer = EventListSerializer(events, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def my_events(self, request):
        """
        Obtener eventos del usuario autenticado.
        """
        if not request.user.is_authenticated:
            return Response(
                {'error': 'Autenticación requerida'},
                status=status.HTTP_401_UNAUTHORIZED
            )
        
        events = self.queryset.filter(organizer=request.user)
        serializer = EventListSerializer(events, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def export_excel(self, request, pk=None):
        """
        Exportar reporte completo del evento a Excel.
        """
        event = self.get_object()
        
        # Crear workbook
        wb = Workbook()

        # Hoja 1: Información del evento
        ws1 = wb.active
        ws1.title = "Información del Evento"

        # Headers con estilo
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        ws1['A1'] = 'EventHub - Reporte de Evento'
        ws1['A1'].font = Font(bold=True, size=16)
        ws1.merge_cells('A1:D1')

        ws1['A3'] = 'Campo'
        ws1['B3'] = 'Valor'
        ws1['A3'].fill = header_fill
        ws1['B3'].fill = header_fill
        ws1['A3'].font = header_font
        ws1['B3'].font = header_font

        # Datos del evento
        info_data = [
            ['Título', event.title],
            ['Categoría', event.category.name],
            ['Lugar', event.venue.name if event.venue else 'N/A'],
            ['Fecha Inicio', event.start_date.strftime('%Y-%m-%d %H:%M')],
            ['Fecha Fin', event.end_date.strftime('%Y-%m-%d %H:%M')],
            ['Capacidad', event.capacity],
            ['Tickets Vendidos', event.tickets_sold],
            ['Tickets Disponibles', event.tickets_available],
            ['Estado', event.get_status_display()],
        ]

        for idx, (field, value) in enumerate(info_data, start=4):
            ws1[f'A{idx}'] = field
            ws1[f'B{idx}'] = value

        # Hoja 2: Tickets vendidos
        ws2 = wb.create_sheet("Tickets Vendidos")
        headers = ['Código', 'Tipo', 'Comprador', 'Email', 'Precio', 'Estado', 'Fecha Compra']

        for col, header in enumerate(headers, start=1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        from apps.tickets.models import Ticket
        tickets = Ticket.objects.filter(
            ticket_type__event=event
        ).select_related('buyer', 'ticket_type')

        for row, ticket in enumerate(tickets, start=2):
            ws2.cell(row=row, column=1, value=ticket.code)
            ws2.cell(row=row, column=2, value=ticket.ticket_type.name)
            ws2.cell(row=row, column=3, value=ticket.buyer.username)
            ws2.cell(row=row, column=4, value=ticket.attendee_email)
            ws2.cell(row=row, column=5, value=float(ticket.final_price))
            ws2.cell(row=row, column=6, value=ticket.get_status_display())
            ws2.cell(row=row, column=7, value=ticket.purchased_at.strftime('%Y-%m-%d %H:%M'))

        # Ajustar ancho de columnas
        for ws in [ws1, ws2]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Generar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=reporte_{event.slug}.xlsx'
        wb.save(response)

        return response
